import imaplib
import email
from email.header import decode_header
import logging
import io
import re
from datetime import datetime
import openpyxl

_LOGGER = logging.getLogger(__name__)

def _try_extract_from_winmail(data: bytes):
    """Try to extract xlsx files from a winmail.dat (TNEF) attachment."""
    try:
        import tnefparse
        tnef = tnefparse.TNEF(data)
        for attachment in tnef.attachments:
            name = getattr(attachment, "name", "") or ""
            if isinstance(name, bytes):
                name = name.decode("utf-8", errors="replace")
            if name.lower().endswith(".xlsx") or name.lower().endswith(".xls"):
                _LOGGER.info(f"Extracted '{name}' from winmail.dat (TNEF)")
                return io.BytesIO(attachment.data)
    except ImportError:
        _LOGGER.warning("tnefparse not installed. Cannot decode winmail.dat.")
    except Exception as e:
        _LOGGER.warning(f"winmail.dat extraction failed: {e}")
    return None

def _decode_str(s):
    """Decode email header string."""
    try:
        if not s:
            return ""
        decoded_list = decode_header(s)
        parts = []
        for content, encoding in decoded_list:
            if isinstance(content, bytes):
                if encoding:
                    try:
                        parts.append(content.decode(encoding))
                    except Exception:
                        parts.append(content.decode("utf-8", errors="replace"))
                else:
                    parts.append(content.decode("utf-8", errors="replace"))
            else:
                parts.append(str(content))
        return "".join(parts)
    except Exception as e:
        _LOGGER.warning(f"Error decoding string: {e}")
        return str(s)

def fetch_from_email(host, port, user, password, subject_filter):
    """Sync function to fetch and process email attachment."""
    try:
        mail = imaplib.IMAP4_SSL(host, port)
        mail.login(user, password)
        mail.select("inbox")

        try:
            typ, data = mail.search(None, 'CHARSET', 'UTF-8', f'(SUBJECT "{subject_filter}")')
        except Exception:
            typ, data = mail.search(None, f'(SUBJECT "{subject_filter}")')

        if not data or not data[0]:
            _LOGGER.info(f"No emails found with subject containing: {subject_filter}")
            mail.close()
            mail.logout()
            return []

        msg_ids = data[0].split()
        msg_ids.reverse()

        rows = []
        found_attachment = False

        for mid in msg_ids[:15]:
            typ, hdr_data = mail.fetch(mid, '(BODY.PEEK[HEADER.FIELDS (SUBJECT DATE)])')
            if not hdr_data or not hdr_data[0]:
                continue

            hdr_msg = email.message_from_bytes(hdr_data[0][1])
            subject_decoded = _decode_str(hdr_msg["Subject"])

            if subject_filter.lower() not in subject_decoded.lower():
                _LOGGER.debug(f"Skipping email '{subject_decoded}' - filter not matched.")
                continue

            typ, msg_data = mail.fetch(mid, '(RFC822)')
            if not msg_data or not msg_data[0]:
                continue

            raw_email = msg_data[0][1]
            msg = email.message_from_bytes(raw_email)
            _LOGGER.info(f"Processing email: '{subject_decoded}' from {msg['Date']}")

            for part in msg.walk():
                if part.get_content_maintype() == 'multipart':
                    continue
                if part.get('Content-Disposition') is None:
                    continue

                filename = part.get_filename()
                if not filename:
                    continue
                filename = _decode_str(filename)

                if filename.lower() == "winmail.dat":
                    _LOGGER.info("Found winmail.dat - attempting TNEF extraction")
                    try:
                        file_data = part.get_payload(decode=True)
                        excel_file = _try_extract_from_winmail(file_data)
                        if excel_file:
                            excel_rows = _parse_excel(excel_file)
                            if excel_rows:
                                rows.extend(excel_rows)
                                found_attachment = True
                                break
                        else:
                            _LOGGER.warning("winmail.dat did not contain a recognizable xlsx.")
                    except Exception as e:
                        _LOGGER.error(f"Error processing winmail.dat: {e}")
                    continue

                if filename.lower().endswith(".xlsx") or filename.lower().endswith(".xls"):
                    _LOGGER.info(f"Found attachment: {filename}")
                    try:
                        file_data = part.get_payload(decode=True)
                        excel_rows = _parse_excel(io.BytesIO(file_data))
                        if excel_rows:
                            rows.extend(excel_rows)
                            found_attachment = True
                            break
                    except Exception as e:
                        _LOGGER.error(f"Error processing attachment {filename}: {e}")

            if found_attachment:
                break

        mail.close()
        mail.logout()
        return rows

    except Exception as e:
        _LOGGER.error(f"IMAP Error: {e}")
        raise Exception(f"IMAP Hiba: {e}") from e


def _parse_excel(file_obj):
    """
    Auto-detect and parse Excel format.

    Új formátum (adatok.xlsx):
      Col0: Pod azonosító
      Col1: Időbélyeg (datetime)
      Col2,3,4: Változó, Érték, Mértékegység (+A / kWh)
      Col5,6,7: Változó, Érték, Mértékegység (+R / ...)
      ... (minden mért változó saját 3 oszlopos csoportban)

    Régi formátum (MertErtekek_*.xlsx):
      Col0: Időbélyeg (datetime vagy szöveges)
      Col1: +A értéke (Num1)
      Col2: -A értéke (Num2)
      ...
    """
    rows = []
    try:
        wb = openpyxl.load_workbook(file_obj, data_only=True)
        ws = wb.active

        if ws.max_row < 2:
            _LOGGER.warning("Excel file has too few rows")
            return rows

        # Read first two header cells to detect format
        first_row = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
        second_row = [cell.value for cell in next(ws.iter_rows(min_row=2, max_row=2))]

        # New format detection:
        # - first_row[0] is "Pod" (or garbled encoding of "Pod")
        # - second_row[1] is a datetime object
        # - first_row[2] contains "ltoz" (Változó, may be garbled)
        is_new_format = (
            ws.max_column >= 5
            and second_row[1] is not None
            and isinstance(second_row[1], datetime)
            and second_row[0] is not None
            and str(second_row[0]).startswith("HU")  # POD identifier
        )

        if is_new_format:
            _LOGGER.info(f"Detected NEW Excel format (Pod|Időbélyeg|Változó|Érték|ME) — {ws.max_row-1} data rows, {ws.max_column} cols")
            rows = _parse_excel_new_format(ws)
        else:
            _LOGGER.info(f"Detected OLD Excel format (Timestamp|+A|-A|...) — {ws.max_row} rows")
            rows = _parse_excel_old_format(ws)

    except Exception as e:
        _LOGGER.error(f"Excel Parse Error: {e}")

    return rows


def _parse_excel_new_format(ws):
    """
    Parse new E.ON Excel format.

    Structure per row:
      [0] Pod  [1] Időbélyeg  [2] Változó  [3] Érték  [4] ME  [5] Változó  [6] Érték  [7] ME  ...

    Each triplet (col i, i+1, i+2) starting at col index 2 contains:
      - the variable name (e.g. '+A', '-A', '1-1:1.8.0*0', ...)
      - the measured value (float or None)
      - the unit of measurement (e.g. 'kWh' or None)
    """
    rows = []

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row or row[0] is None or row[1] is None:
            continue

        pod = str(row[0]).strip()
        ts_cell = row[1]

        if not isinstance(ts_cell, datetime):
            continue

        ts_ms = int(ts_cell.timestamp() * 1000)

        mapped_row = {
            "Timestamp": f"/Date({ts_ms})/",
            "Datum": ts_cell.strftime("%Y-%m-%d"),
            "Pod": pod,
            "Num1": 0.0,  # +A import
            "Num2": 0.0,  # -A export
        }

        # Parse all variable triplets: columns 2..end, step 3
        col_idx = 2
        num_cols = len(row)
        while col_idx + 1 < num_cols:
            var_name = row[col_idx]
            raw_val  = row[col_idx + 1]
            # unit   = row[col_idx + 2]  # available but not stored separately

            if var_name is not None:
                var_name = str(var_name).strip()
                try:
                    value = float(raw_val) if raw_val is not None else 0.0
                except (ValueError, TypeError):
                    value = 0.0

                # Store under the original OBIS / variable name
                mapped_row[var_name] = value

                # Backward-compatible aliases used by coordinator & sensors
                if var_name == "+A":
                    mapped_row["Num1"] = value
                elif var_name == "-A":
                    mapped_row["Num2"] = value

            col_idx += 3

        rows.append(mapped_row)

    _LOGGER.info(f"New-format parse: {len(rows)} rows extracted, first Pod: {rows[0]['Pod'] if rows else 'N/A'}")
    return rows


def _parse_excel_old_format(ws):
    """
    Parse legacy E.ON Excel format (MertErtekek_*.xlsx).

    Structure:
      Col0: Timestamp (datetime object OR '2024.03.01. 00:00' string)
      Col1: +A value  → Num1
      Col2: -A value  → Num2
      Col3+: extra columns stored under their header name
    """
    rows = []
    date_pattern = re.compile(r"^\d{4}\.\d{2}\.\d{2}\. \d{2}:\d{2}$")

    headers = []
    for i, col in enumerate(ws.iter_cols(1, ws.max_column, 1, 1, values_only=True)):
        head_val = col[0]
        headers.append(str(head_val).strip() if head_val else f"Col_{i}")

    for row in ws.iter_rows(values_only=True):
        if not row or len(row) < 3:
            continue

        col_a = row[0]
        dt = None

        if isinstance(col_a, datetime):
            dt = col_a
        else:
            s_val = str(col_a).strip() if col_a else ""
            if date_pattern.match(s_val):
                try:
                    dt = datetime.strptime(s_val, "%Y.%m.%d. %H:%M")
                except ValueError:
                    pass

        if not dt:
            continue

        ts_ms = int(dt.timestamp() * 1000)

        mapped_row = {
            "Timestamp": f"/Date({ts_ms})/",
            "Datum": dt.strftime("%Y-%m-%d"),
            "Pod": "EmailData",
            "Num1": 0.0,
            "Num2": 0.0,
        }

        for idx, val in enumerate(row):
            if idx == 0:
                continue
            col_name = headers[idx] if idx < len(headers) else f"Col_{idx}"
            safe_val = float(val) if val is not None else 0.0
            if idx == 1:
                mapped_row["Num1"] = safe_val
                mapped_row["+A"] = safe_val
            elif idx == 2:
                mapped_row["Num2"] = safe_val
                mapped_row["-A"] = safe_val
            else:
                mapped_row[col_name] = safe_val

        rows.append(mapped_row)

    return rows


